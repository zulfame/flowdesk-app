import io
from datetime import datetime, timedelta

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List

from db import db
from helpers import new_id, now_iso, log_activity, is_privileged, can_manage, schedule_scope_query, scope_user_ids
from security import get_current_user
from services import delete_time_schedule

router = APIRouter(prefix="/time-schedules", tags=["time-schedules"])

CATEGORIES = {"pelaksanaan", "event", "libur"}
# Selaras dengan kartu Linimasa (monokrom bila kegiatan tidak punya warna sendiri).
CAT_FILL = {"pelaksanaan": "F0A21B", "event": "10B27A", "libur": "E8546F"}
MONTHS_ID = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]


class Person(BaseModel):
    user_id: Optional[str] = None
    name: Optional[str] = ""
    department: Optional[str] = ""
    email: Optional[str] = ""
    phone: Optional[str] = ""


class Activity(BaseModel):
    id: Optional[str] = None
    name: str
    section: Optional[str] = ""
    pic: Optional[Person] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    category: str = "pelaksanaan"
    color: Optional[str] = None
    status: str = "Rencana"
    note: Optional[str] = ""
    task_id: Optional[str] = None


class ScheduleCreate(BaseModel):
    title: str
    event_name: Optional[str] = ""
    section: Optional[str] = ""
    description: Optional[str] = ""
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    holidays: List[str] = []
    event_dates: List[str] = []
    activities: List[Activity] = []


class ScheduleUpdate(BaseModel):
    title: Optional[str] = None
    event_name: Optional[str] = None
    section: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    holidays: Optional[List[str]] = None
    event_dates: Optional[List[str]] = None
    activities: Optional[List[Activity]] = None


class ConvertBody(BaseModel):
    pic: Optional[Person] = None
    priority: str = "Medium"
    deadline: Optional[str] = None


async def _vis(user: dict) -> dict:
    return await schedule_scope_query(db, user)


async def _can_view(user: dict, s: dict) -> bool:
    uid = user.get("id")
    if s.get("created_by") == uid:
        return True
    if any((a.get("pic") or {}).get("user_id") == uid for a in s.get("activities", [])):
        return True
    ids = await scope_user_ids(db, user)
    if ids is None:
        return True
    return s.get("created_by") in ids or any(
        (a.get("pic") or {}).get("user_id") in ids for a in s.get("activities", [])
    )


def _norm_activities(acts) -> list:
    out = []
    for a in acts or []:
        d = a.model_dump() if hasattr(a, "model_dump") else dict(a)
        if not d.get("id"):
            d["id"] = new_id()
        if d.get("category") not in CATEGORIES:
            d["category"] = "pelaksanaan"
        out.append(d)
    return out


@router.get("")
async def list_schedules(user: dict = Depends(get_current_user)):
    return await db.time_schedules.find(
        {**await _vis(user), "is_deleted": {"$ne": True}}, {"_id": 0}
    ).sort("created_at", -1).to_list(500)


@router.post("")
async def create_schedule(body: ScheduleCreate, user: dict = Depends(get_current_user)):
    doc = body.model_dump()
    doc["activities"] = _norm_activities(body.activities)
    doc.update({
        "id": new_id(), "is_deleted": False,
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": now_iso(), "updated_at": now_iso(),
    })
    await db.time_schedules.insert_one(dict(doc))
    doc.pop("_id", None)
    await log_activity(db, user, "create", "time_schedule", doc["id"], f"Membuat jadwal '{doc['title']}'")
    return doc


@router.get("/{sid}")
async def get_schedule(sid: str, user: dict = Depends(get_current_user)):
    s = await db.time_schedules.find_one({"id": sid, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")
    if not await _can_view(user, s):
        raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke jadwal ini")
    return s


@router.put("/{sid}")
async def update_schedule(sid: str, body: ScheduleUpdate, user: dict = Depends(get_current_user)):
    existing = await db.time_schedules.find_one({"id": sid, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")
    if not can_manage(user, existing):
        raise HTTPException(status_code=403, detail="Hanya pembuat jadwal atau Admin yang dapat mengubah")
    update = {k: v for k, v in body.model_dump().items() if v is not None}
    if "activities" in update:
        update["activities"] = _norm_activities(body.activities)
    update["updated_at"] = now_iso()
    await db.time_schedules.update_one({"id": sid}, {"$set": update})
    await log_activity(db, user, "update", "time_schedule", sid, f"Memperbarui jadwal '{existing['title']}'")
    return await db.time_schedules.find_one({"id": sid}, {"_id": 0})


@router.delete("/{sid}")
async def delete_schedule(sid: str, user: dict = Depends(get_current_user)):
    existing = await db.time_schedules.find_one({"id": sid, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")
    if not can_manage(user, existing):
        raise HTTPException(status_code=403, detail="Hanya pembuat jadwal atau Admin yang dapat menghapus")
    await delete_time_schedule(sid, user)
    return {"message": "Jadwal dihapus"}


@router.post("/{sid}/activities/{aid}/convert-task")
async def convert_activity(sid: str, aid: str, body: ConvertBody, user: dict = Depends(get_current_user)):
    s = await db.time_schedules.find_one({"id": sid, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")
    if not await _can_view(user, s):
        raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke jadwal ini")
    act = next((a for a in s.get("activities", []) if a.get("id") == aid), None)
    if not act:
        raise HTTPException(status_code=404, detail="Kegiatan tidak ditemukan")
    if act.get("task_id"):
        raise HTTPException(status_code=400, detail="Kegiatan ini sudah memiliki tugas terkait")

    from routers.tasks import compute, EMPTY_PERSON, _notify_pic
    pic = body.pic.model_dump() if body.pic else (act.get("pic") or dict(EMPTY_PERSON))
    if not pic or not pic.get("name"):
        raise HTTPException(status_code=400, detail="PIC wajib dipilih untuk membuat tugas")
    deadline = body.deadline or act.get("end_date")

    task = {
        "id": new_id(),
        "title": act["name"],
        "description": act.get("note") or f"Dibuat dari jadwal '{s['title']}'",
        "requester": dict(EMPTY_PERSON),
        "pic": {**dict(EMPTY_PERSON), **pic},
        "priority": body.priority,
        "deadline": deadline,
        "items": [], "documents": [], "comments": [],
        "status": "Pending",
        "time_schedule_id": sid,
        "history": [{"action": "created", "by": user["name"], "at": now_iso(),
                     "detail": f"Dari kegiatan jadwal '{s['title']}'"}],
        "created_by": user["id"], "created_by_name": user["name"],
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    task = compute(task)
    await db.tasks.insert_one(dict(task))
    task.pop("_id", None)
    await db.time_schedules.update_one(
        {"id": sid, "activities.id": aid}, {"$set": {"activities.$.task_id": task["id"]}}
    )
    await log_activity(db, user, "create", "task", task["id"], f"Membuat tugas dari kegiatan '{act['name']}'")
    try:
        await _notify_pic(task)
    except Exception:
        pass
    return {"task_id": task["id"], "title": task["title"]}


def _daterange(start: str, end: str):
    try:
        d = datetime.fromisoformat(start[:10])
        e = datetime.fromisoformat(end[:10])
    except Exception:
        return []
    if e < d or (e - d).days > 500:
        return []
    out = []
    while d <= e:
        out.append(d)
        d += timedelta(days=1)
    return out


@router.get("/{sid}/export")
async def export_schedule(sid: str, user: dict = Depends(get_current_user)):
    s = await db.time_schedules.find_one({"id": sid, "is_deleted": {"$ne": True}}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Jadwal tidak ditemukan")
    if not await _can_view(user, s):
        raise HTTPException(status_code=403, detail="Anda tidak memiliki akses ke jadwal ini")

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    activities = s.get("activities", [])
    start = s.get("start_date")
    end = s.get("end_date")
    if not start or not end:
        ds = [a.get("start_date") for a in activities if a.get("start_date")]
        de = [a.get("end_date") for a in activities if a.get("end_date")]
        start = start or (min(ds) if ds else datetime.now().date().isoformat())
        end = end or (max(de) if de else start)
    days = _daterange(start, end)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Linimasa"
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="F3F4F6")
    holiday_fill = PatternFill("solid", fgColor="FEF1F2")
    holiday_font = "D33A57"
    event_fill = PatternFill("solid", fgColor="E5E7EB")
    center = Alignment(horizontal="center", vertical="center")
    today_side = Side(style="medium", color="111827")

    holidays = set(s.get("holidays") or [])
    event_dates = {str(d)[:10] for d in (s.get("event_dates") or [])}
    today_key = datetime.now().date().isoformat()
    day_keys = [d.date().isoformat() for d in days]

    ws.cell(row=1, column=1, value=s.get("title") or "Linimasa").font = Font(bold=True, size=14)
    if s.get("event_name"):
        ws.cell(row=2, column=1, value=s["event_name"]).font = Font(size=10, color="6B7280")

    MONTH_ROW, DAY_ROW, FIRST_COL = 4, 5, 3

    # Baris bulan: gabung kolom per bulan, seperti header "Agu 2026" di kartu.
    groups = []
    for i, d in enumerate(days):
        key = (d.year, d.month)
        if groups and groups[-1][0] == key:
            groups[-1][2] = i
        else:
            groups.append([key, i, i])
    for (year, month), i0, i1 in groups:
        c0, c1 = FIRST_COL + i0, FIRST_COL + i1
        if c1 > c0:
            ws.merge_cells(start_row=MONTH_ROW, start_column=c0, end_row=MONTH_ROW, end_column=c1)
        cell = ws.cell(row=MONTH_ROW, column=c0, value=f"{MONTHS_ID[month - 1]} {year}")
        cell.font = Font(bold=True, size=9)
        cell.alignment = center
        for c in range(c0, c1 + 1):
            ws.cell(row=MONTH_ROW, column=c).fill = head_fill
            ws.cell(row=MONTH_ROW, column=c).border = border

    ws.merge_cells(start_row=MONTH_ROW, start_column=1, end_row=DAY_ROW, end_column=1)
    ws.merge_cells(start_row=MONTH_ROW, start_column=2, end_row=DAY_ROW, end_column=2)
    for col, label in ((1, "KEGIATAN"), (2, "PIC")):
        cell = ws.cell(row=MONTH_ROW, column=col, value=label)
        cell.font = Font(bold=True, size=9, color="6B7280")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = head_fill
        cell.border = border

    for i, d in enumerate(days):
        key = day_keys[i]
        weekend = d.weekday() >= 5
        cell = ws.cell(row=DAY_ROW, column=FIRST_COL + i, value=d.day)
        holiday = weekend or key in holidays
        cell.font = Font(bold=key == today_key or holiday, size=8,
                         color=holiday_font if holiday else "6B7280")
        cell.alignment = center
        cell.border = border
        cell.fill = (holiday_fill if holiday
                     else event_fill if key in event_dates else head_fill)
        ws.column_dimensions[cell.column_letter].width = 3.6

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = ws.cell(row=DAY_ROW + 1, column=FIRST_COL)

    for r, a in enumerate(activities, start=DAY_ROW + 1):
        name = ws.cell(row=r, column=1, value=a.get("name", ""))
        name.font = Font(bold=True, size=10)
        name.border = border
        pic = ws.cell(row=r, column=2, value=(a.get("pic") or {}).get("name") or "Tanpa PIC")
        pic.font = Font(size=9, color="6B7280")
        pic.border = border

        raw_color = (a.get("color") or "").lstrip("#")
        hexc = raw_color.upper() if len(raw_color) == 6 else CAT_FILL.get(a.get("category", "pelaksanaan"), "5B5B5B")
        bar = PatternFill("solid", fgColor=hexc)
        a_start = (a.get("start_date") or "")[:10]
        a_end = (a.get("end_date") or "")[:10]
        for i, dk in enumerate(day_keys):
            cell = ws.cell(row=r, column=FIRST_COL + i)
            cell.border = border
            weekend = days[i].weekday() >= 5
            if a_start and a_end and a_start <= dk <= a_end:
                cell.fill = bar
            elif weekend or dk in holidays:
                cell.fill = holiday_fill
            elif dk in event_dates:
                cell.fill = event_fill
            if dk == today_key:
                cell.border = Border(left=today_side, right=thin, top=thin, bottom=thin)
        ws.row_dimensions[r].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = (s.get("title") or "linimasa").replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
    )
