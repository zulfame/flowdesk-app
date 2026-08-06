from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel, EmailStr, Field
from typing import Optional, List, Dict
import io
import csv
import re

from db import db
from helpers import new_id, now_iso, log_activity, ROLE_LEVELS, guess_role_level, subordinate_users
from security import hash_password, get_current_user, require_admin

router = APIRouter(tags=["users"])

DEFAULT_IMPORT_PASSWORD = "flowdesk123"

PERMISSION_CATALOG = [
    {"key": "task", "label": "Kelola Tugas"},
    {"key": "meeting", "label": "Kelola Rapat"},
    {"key": "time_schedule", "label": "Time Schedule"},
    {"key": "note", "label": "Kelola Catatan"},
    {"key": "reminder", "label": "Ingatkan Saya"},
    {"key": "calendar", "label": "Kalender"},
    {"key": "help_ticket", "label": "Tiket Bantuan"},
    {"key": "report", "label": "Laporan & Ekspor"},
    {"key": "user", "label": "Kelola Pengguna"},
    {"key": "role", "label": "Kelola Peranan"},
    {"key": "database", "label": "Kelola Database"},
    {"key": "notification_config", "label": "Kelola Notifikasi"},
    {"key": "app_config", "label": "Kelola Aplikasi"},
    {"key": "activity", "label": "Log Aktivitas"},
]

DEFAULT_ROLES = [
    {"name": "super_admin", "label": "Super Admin", "permissions": ["*"], "parent_id": None,
     "level": "Dirut", "is_system": True},
    {"name": "guest", "label": "Guest", "permissions": [], "parent_id": None,
     "level": "Staff", "is_system": True},
]


def _slug(label: str) -> str:
    return re.sub(r"^_+|_+$", "", re.sub(r"[^a-z0-9]+", "_", (label or "").lower()))


class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=6)
    role: str = "guest"
    phone: Optional[str] = None
    department: Optional[str] = None


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    department: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None


class RoleBody(BaseModel):
    name: str
    label: str
    permissions: List[str] = []
    parent_id: Optional[str] = None
    level: Optional[str] = None


def _clean(u: dict) -> dict:
    u.pop("password_hash", None)
    u.pop("_id", None)
    return u


@router.get("/users")
async def list_users(page: int = 1, page_size: int = 20, q: Optional[str] = None,
                     role: Optional[str] = None, all: bool = False,
                     user: dict = Depends(get_current_user)):
    query = {}
    if q:
        rx = {"$regex": q, "$options": "i"}
        query["$or"] = [{"name": rx}, {"email": rx}, {"department": rx}]
    if role and role != "all":
        query["role"] = role
    proj = {"_id": 0, "password_hash": 0}
    if all:
        items = await db.users.find(query, proj).sort("created_at", -1).to_list(5000)
        return {"items": items, "total": len(items), "page": 1, "page_size": len(items)}
    total = await db.users.count_documents(query)
    page = max(1, page)
    page_size = min(max(1, page_size), 100)
    items = await db.users.find(query, proj).sort("created_at", -1) \
        .skip((page - 1) * page_size).limit(page_size).to_list(page_size)
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/users/subordinates")
async def list_subordinates(user: dict = Depends(get_current_user)):
    """Kandidat PIC: pemegang jabatan DI BAWAH jabatan pengguna (semua turunan)."""
    items = await subordinate_users(db, user)
    items.sort(key=lambda u: (u.get("name") or "").lower())
    return {"items": items, "total": len(items)}


@router.get("/permissions")
async def list_permissions(user: dict = Depends(get_current_user)):
    return PERMISSION_CATALOG


@router.post("/users/import")
async def import_users(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    raw = await file.read()
    fname = (file.filename or "").lower()
    rows = []
    try:
        if fname.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            if data:
                headers = [str(h).strip().lower() if h is not None else "" for h in data[0]]
                for r in data[1:]:
                    rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})
        else:
            text = raw.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca berkas: {e}")

    created, updated, errors = 0, 0, []
    valid_roles = {r["name"] for r in (await db.roles.find({}, {"_id": 0, "name": 1}).to_list(100))} or {"super_admin", "guest"}

    for idx, row in enumerate(rows, start=2):
        name = str(row.get("name") or row.get("nama") or "").strip()
        email = str(row.get("email") or "").strip().lower()
        if not email or "@" not in email:
            if name or email:
                errors.append(f"Baris {idx}: email tidak valid")
            continue
        role = str(row.get("role") or row.get("peran") or "guest").strip().lower()
        if role not in valid_roles:
            role = "guest"
        phone = str(row.get("phone") or row.get("telepon") or "").strip() or None
        dept = str(row.get("department") or row.get("departemen") or "").strip() or None
        fields = {"name": name or email.split("@")[0], "role": role, "phone": phone, "department": dept}
        existing = await db.users.find_one({"email": email})
        if existing:
            await db.users.update_one({"email": email}, {"$set": fields})
            updated += 1
        else:
            password = str(row.get("password") or "").strip() or DEFAULT_IMPORT_PASSWORD
            await db.users.insert_one({
                "id": new_id(), "email": email, "password_hash": hash_password(password),
                "permissions": [], "avatar": None, "is_active": True, "created_at": now_iso(),
                **fields,
            })
            created += 1

    await log_activity(db, admin, "create", "user", None, f"Impor pengguna: {created} baru, {updated} diperbarui")
    return {"created": created, "updated": updated, "errors": errors,
            "default_password": DEFAULT_IMPORT_PASSWORD}


@router.post("/users")
async def create_user(body: UserCreate, admin: dict = Depends(require_admin)):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    doc = {
        "id": new_id(),
        "name": body.name,
        "email": email,
        "password_hash": hash_password(body.password),
        "role": body.role,
        "permissions": [],
        "phone": body.phone,
        "department": body.department,
        "avatar": None,
        "is_active": True,
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    await log_activity(db, admin, "create", "user", doc["id"], f"Membuat pengguna {email}")
    return _clean(dict(doc))


@router.put("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdate, admin: dict = Depends(require_admin)):
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    update = {k: v for k, v in body.model_dump().items() if v is not None and k != "password"}
    if body.password:
        update["password_hash"] = hash_password(body.password)
    await db.users.update_one({"id": user_id}, {"$set": update})
    await log_activity(db, admin, "update", "user", user_id, f"Memperbarui pengguna {existing['email']}")
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated


@router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun sendiri")
    existing = await db.users.find_one({"id": user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Pengguna tidak ditemukan")
    await db.users.delete_one({"id": user_id})
    # Cascade: keluarkan dari rapat & hapus data rapat pribadinya (entry + lampiran)
    uname = existing.get("name")
    async for m in db.meetings.find({"member_ids": user_id}, {"_id": 0, "id": 1, "participants": 1}):
        await db.meetings.update_one(
            {"id": m["id"]},
            {"$pull": {"member_ids": user_id, "participants": uname}, "$unset": {f"entries.{user_id}": ""}},
        )
        await db.files.delete_many({"parent_id": f"{m['id']}:{user_id}"})
    await log_activity(db, admin, "delete", "user", user_id, f"Menghapus pengguna {existing['email']}")
    return {"message": "Pengguna dihapus"}


def _sort_roles(roles: list) -> list:
    """Urut hierarkis: induk lalu turunannya, tiap tingkat urut nama jabatan."""
    children = {}
    for r in roles:
        children.setdefault(r.get("parent_id") or None, []).append(r)
    for v in children.values():
        v.sort(key=lambda r: (r.get("label") or "").lower())
    out = []

    def walk(pid, depth):
        for r in children.get(pid, []):
            r["depth"] = depth
            out.append(r)
            walk(r["id"], depth + 1)

    walk(None, 0)
    seen = {r["id"] for r in out}
    for r in roles:  # peran dengan induk yatim tetap ditampilkan
        if r["id"] not in seen:
            r["depth"] = 0
            out.append(r)
    return out


@router.get("/roles")
async def list_roles(user: dict = Depends(get_current_user)):
    roles = await db.roles.find({}, {"_id": 0}).to_list(500)
    if not roles:
        for r in DEFAULT_ROLES:
            await db.roles.insert_one({"id": new_id(), **r})
        roles = await db.roles.find({}, {"_id": 0}).to_list(500)
    by_id = {r["id"]: r for r in roles}
    for r in roles:
        r.setdefault("parent_id", None)
        r.setdefault("is_system", False)
        r["level"] = r.get("level") or guess_role_level(r.get("label"))
        parent = by_id.get(r.get("parent_id"))
        r["parent_label"] = parent.get("label") if parent else None
    return _sort_roles(roles)


@router.get("/role-levels")
async def get_role_levels(user: dict = Depends(get_current_user)):
    """Izin bawaan per level jabatan — diwarisi peran yang izinnya belum ditimpa."""
    doc = await db.settings.find_one({"key": "app"}, {"_id": 0, "role_levels": 1})
    saved = (doc or {}).get("role_levels") or {}
    return {"levels": ROLE_LEVELS, "permissions": {lv: saved.get(lv, []) for lv in ROLE_LEVELS}}


class RoleLevelsBody(BaseModel):
    permissions: Dict[str, List[str]]


@router.put("/role-levels")
async def set_role_levels(body: RoleLevelsBody, admin: dict = Depends(require_admin)):
    clean = {lv: sorted(set(body.permissions.get(lv) or [])) for lv in ROLE_LEVELS}
    await db.settings.update_one({"key": "app"}, {"$set": {"role_levels": clean}}, upsert=True)
    await log_activity(db, admin, "update", "role", None, "Memperbarui izin per level jabatan")
    return {"permissions": clean}


@router.post("/roles/import")
async def import_roles(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    """Impor hierarki jabatan dari Excel/CSV: kolom Name, Parent (atau Parent ID), Level, Order."""
    raw = await file.read()
    fname = (file.filename or "").lower()
    rows = []
    try:
        if fname.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            data = list(wb.active.iter_rows(values_only=True))
            if data:
                headers = [str(h).strip().lower() if h is not None else "" for h in data[0]]
                rows = [{headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))} for r in data[1:]]
        else:
            text = raw.decode("utf-8-sig")
            rows = [{(k or "").strip().lower(): v for k, v in row.items()} for row in csv.DictReader(io.StringIO(text))]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca berkas: {e}")

    parsed = []
    for row in rows:
        label = str(row.get("name") or row.get("nama") or row.get("label") or "").strip()
        if not label:
            continue
        parsed.append({
            "ext_id": str(row.get("id") or "").strip() or None,
            "label": label,
            "parent_label": str(row.get("parent") or row.get("atasan") or "").strip() or None,
            "parent_ext_id": str(row.get("parent id") or row.get("parent_id") or "").strip() or None,
            "level": str(row.get("level") or "").strip() or guess_role_level(label),
        })
    if not parsed:
        raise HTTPException(status_code=400, detail="Tidak ada baris jabatan yang bisa dibaca")

    created, updated = 0, 0
    id_by_ext, id_by_label = {}, {}
    for r in parsed:
        name = _slug(r["label"])
        existing = await db.roles.find_one({"name": name}, {"_id": 0})
        fields = {"label": r["label"], "level": r["level"]}
        if existing:
            await db.roles.update_one({"name": name}, {"$set": fields})
            rid = existing["id"]
            updated += 1
        else:
            rid = new_id()
            system = name in ("super_admin", "guest")
            await db.roles.insert_one({"id": rid, "name": name,
                                       "permissions": ["*"] if name == "super_admin" else [],
                                       "parent_id": None, "is_system": system, **fields})
            created += 1
        if r["ext_id"]:
            id_by_ext[r["ext_id"]] = rid
        id_by_label[r["label"].lower()] = rid

    linked = 0
    for r in parsed:
        rid = id_by_ext.get(r["ext_id"]) or id_by_label.get(r["label"].lower())
        parent = None
        if r["parent_ext_id"]:
            parent = id_by_ext.get(r["parent_ext_id"])
        if not parent and r["parent_label"]:
            parent = id_by_label.get(r["parent_label"].lower())
        if parent and parent != rid:
            await db.roles.update_one({"id": rid}, {"$set": {"parent_id": parent}})
            linked += 1

    await log_activity(db, admin, "create", "role", None, f"Impor {created + updated} jabatan dari berkas")
    return {"created": created, "updated": updated, "linked": linked, "total": len(parsed)}


@router.post("/roles")
async def create_role(body: RoleBody, admin: dict = Depends(require_admin)):
    if await db.roles.find_one({"name": body.name}):
        raise HTTPException(status_code=400, detail="Role sudah ada")
    doc = {"id": new_id(), **body.model_dump()}
    await db.roles.insert_one(doc)
    doc.pop("_id", None)
    await log_activity(db, admin, "create", "role", doc["id"], f"Membuat role {body.name}")
    return doc


@router.put("/roles/{role_id}")
async def update_role(role_id: str, body: RoleBody, admin: dict = Depends(require_admin)):
    await db.roles.update_one({"id": role_id}, {"$set": body.model_dump()})
    return await db.roles.find_one({"id": role_id}, {"_id": 0})


@router.delete("/roles/{role_id}")
async def delete_role(role_id: str, admin: dict = Depends(require_admin)):
    role = await db.roles.find_one({"id": role_id})
    if role and (role.get("is_system") or role.get("name") in ("super_admin", "guest")):
        raise HTTPException(status_code=400, detail="Role bawaan tidak dapat dihapus")
    await db.roles.delete_one({"id": role_id})
    return {"message": "Role dihapus"}
